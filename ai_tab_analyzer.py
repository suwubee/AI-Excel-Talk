#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
AI分析Tab专用分析器 - 增强版
- 智能识别标准表格 vs 复杂表格
- 处理合并单元格和多行表头
- AI提示词友好输出
- 自动识别筛选项字段
"""

import openpyxl
from openpyxl.utils import get_column_letter
from collections import Counter

class AITabAnalyzer:
    """AI分析Tab专用的Excel分析器 - 增强版"""
    
    def analyze_for_ai(self, file_path: str) -> str:
        """
        为AI分析生成智能的Excel结构描述
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            AI友好的markdown格式描述
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            lines = []
            lines.append(f"# 📊 Excel数据结构分析")
            lines.append(f"**文件**: {file_path.split('/')[-1].split('\\')[-1]}")
            lines.append("")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_analysis = self._analyze_sheet_intelligent(ws, sheet_name)
                lines.append(sheet_analysis)
                lines.append("")
                
            return "\n".join(lines)
            
        except Exception as e:
            return f"# ❌ 分析失败\n\n错误信息: {str(e)}"
    
    def _analyze_sheet_intelligent(self, ws, sheet_name: str) -> str:
        """智能分析工作表"""
        # 基本信息
        rows, cols = ws.max_row, ws.max_column
        merged_ranges = list(ws.merged_cells.ranges)
        has_merges = len(merged_ranges) > 0
        
        # 判断表格类型
        is_standard = self._is_standard_table(ws, merged_ranges)
        
        result = []
        result.append(f"## 📋 {sheet_name}")
        result.append(f"**规模**: {rows}行 × {cols}列")
        
        if is_standard:
            result.append(f"**类型**: 🟢标准二维表格")
            analysis = self._analyze_standard_table(ws)
        else:
            result.append(f"**类型**: 🟡复杂表格 (含{len(merged_ranges)}个合并单元格)")
            analysis = self._analyze_complex_table(ws, merged_ranges)
        
        result.extend(analysis)
        return "\n".join(result)
    
    def _is_standard_table(self, ws, merged_ranges) -> bool:
        """判断是否为标准二维表格"""
        # 如果有合并单元格，不是标准表格
        if len(merged_ranges) > 0:
            return False
        
        # 检查前5行，找到最可能的表头行
        best_header_row = self._find_header_by_data_consistency(ws)
        
        # 如果找到了明确的表头行，就是标准表格
        return best_header_row is not None
    
    def _find_header_by_data_consistency(self, ws) -> int:
        """通过数据一致性分析找到表头行"""
        for row in range(1, min(6, ws.max_row + 1)):
            if self._is_likely_header_row(ws, row):
                return row
        return None
    
    def _is_likely_header_row(self, ws, row: int) -> bool:
        """判断某一行是否可能是表头行（基于统计学分析的通用方法）"""
        if row >= ws.max_row:
            return False
            
        # 1. 检查当前行是否有足够的非空单元格
        current_row_data = []
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row, col)
            value = str(cell.value).strip() if cell.value else ""
            current_row_data.append(value)
        
        non_empty_current = sum(1 for v in current_row_data if v)
        if non_empty_current < 3:  # 至少要有3个非空字段
            return False
        
        # 2. 关键：比较当前行与后续数据行的特征差异
        if row + 2 > ws.max_row:  # 至少需要2行数据来做比较
            # 如果没有足够的数据行，用更宽松的判断
            return self._analyze_single_row_header_likelihood(current_row_data) >= 0.6
            
        # 3. 统计学分析：比较当前行与数据行的差异
        header_likelihood = self._analyze_header_vs_data_pattern(ws, row, current_row_data)
        
        return header_likelihood >= 0.5  # 降低阈值，更宽松的判断
    
    def _analyze_single_row_header_likelihood(self, row_data: list) -> float:
        """当没有足够数据行时，分析单行作为表头的可能性"""
        if not row_data:
            return 0.0
        
        scores = []
        for value in row_data:
            if not value.strip():
                continue
                
            score = 0.0
            
            # 长度评分：表头字段通常较短
            if len(value) <= 10:
                score += 0.3
            elif len(value) <= 20:
                score += 0.1
            
            # 复杂度评分：表头字段通常较简单
            if not any(c in value for c in ['。', '，', '；', '：', '？', '！', '.', ',', ';', ':', '?', '!']):
                score += 0.3
            
            # 数字评分：纯长数字通常不是表头
            if self._is_numeric(value) and len(value) > 4:
                score -= 0.2
            
            # 日期评分：日期格式通常不是表头
            if self._looks_like_date(value):
                score -= 0.3
            
            scores.append(max(0.0, min(1.0, score)))  # 限制在0-1之间
        
        return sum(scores) / len(scores) if scores else 0.0
    
    def _analyze_header_vs_data_pattern(self, ws, header_row: int, header_data: list) -> float:
        """分析表头行与数据行的模式差异"""
        # 收集后续数据行
        data_rows = []
        for row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
            row_data = []
            for col in range(1, len(header_data) + 1):
                cell = ws.cell(row, col)
                value = str(cell.value).strip() if cell.value else ""
                row_data.append(value)
            if any(v for v in row_data):  # 只要有非空数据
                data_rows.append(row_data)
        
        if not data_rows:
            return 0.0
        
        # 分析差异得分
        differences = []
        
        for col_idx in range(len(header_data)):
            if not header_data[col_idx]:
                continue
                
            header_value = header_data[col_idx]
            data_values = [row[col_idx] for row in data_rows if col_idx < len(row) and row[col_idx]]
            
            if not data_values:
                continue
            
            # 计算该列表头与数据的差异程度
            diff_score = self._calculate_column_difference(header_value, data_values)
            differences.append(diff_score)
        
        return sum(differences) / len(differences) if differences else 0.0
    
    def _calculate_column_difference(self, header_value: str, data_values: list) -> float:
        """计算表头值与数据值的差异程度"""
        score = 0.0
        
        # 1. 长度差异：表头通常比数据短或相近
        avg_data_len = sum(len(v) for v in data_values) / len(data_values)
        if len(header_value) <= avg_data_len:
            score += 0.3
        
        # 2. 类型差异：表头与数据类型不同得分高
        header_type = self._classify_data_type(header_value)
        data_types = [self._classify_data_type(v) for v in data_values]
        data_type_counts = {}
        for dt in data_types:
            data_type_counts[dt] = data_type_counts.get(dt, 0) + 1
        
        if data_type_counts:
            main_data_type = max(data_type_counts.keys(), key=lambda k: data_type_counts[k])
            if header_type != main_data_type:
                score += 0.4
        
        # 3. 重复性差异：表头通常不重复，数据可能重复
        data_unique_ratio = len(set(data_values)) / len(data_values)
        if data_unique_ratio < 0.8:  # 数据有重复
            score += 0.3
        
        return min(1.0, score)
    
    def _analyze_data_consistency(self, ws, start_row: int, num_cols: int) -> float:
        """分析数据的一致性程度"""
        if start_row > ws.max_row:
            return 0.0
        
        # 检查后续5行的数据类型一致性
        col_types = []  # 每列的数据类型统计
        
        for col in range(1, min(num_cols + 1, 20)):
            types_in_col = []
            
            # 取样5行数据
            for row in range(start_row, min(start_row + 5, ws.max_row + 1)):
                cell = ws.cell(row, col)
                if cell.value is not None:
                    value_str = str(cell.value).strip()
                    if value_str:
                        data_type = self._classify_data_type(value_str)
                        types_in_col.append(data_type)
            
            if types_in_col:
                # 计算该列数据类型的一致性
                type_counts = {}
                for t in types_in_col:
                    type_counts[t] = type_counts.get(t, 0) + 1
                
                # 主要类型占比
                max_count = max(type_counts.values())
                consistency = max_count / len(types_in_col)
                col_types.append(consistency)
        
        # 返回平均一致性
        return sum(col_types) / len(col_types) if col_types else 0.0
    
    def _classify_data_type(self, value_str: str) -> str:
        """分类数据类型"""
        if self._is_numeric(value_str):
            return "numeric"
        elif self._looks_like_date(value_str):
            return "date"
        elif len(value_str) <= 5:
            return "short_text"
        elif len(value_str) <= 20:
            return "medium_text"
        else:
            return "long_text"
    
    def _analyze_header_quality(self, row_data: list) -> float:
        """分析表头质量"""
        if not row_data:
            return 0.0
        
        valid_headers = 0
        total_non_empty = 0
        
        for value in row_data:
            if value.strip():
                total_non_empty += 1
                # 只使用严格的表头检测逻辑，去掉长度兜底
                if self._looks_like_header(value):
                    valid_headers += 1
        
        return valid_headers / total_non_empty if total_non_empty > 0 else 0.0
    
    def _analyze_standard_table(self, ws) -> list:
        """分析标准二维表格"""
        result = []
        
        # 找表头
        header_row = self._find_best_header_row(ws)
        header_cell = f"{get_column_letter(1)}{header_row}"
        result.append(f"**📍 表头位置**: `{header_cell}` (第{header_row}行)")
        
        # 数据起始位置
        data_start_row = header_row + 1
        data_start_cell = f"{get_column_letter(1)}{data_start_row}"
        result.append(f"**📍 数据起始位置**: `{data_start_cell}` (第{data_start_row}行第1列)")
        
        # 提取字段
        fields = self._extract_standard_fields(ws, header_row)
        
        result.append("")
        result.append("**🏷️ 字段结构**:")
        
        for i, field in enumerate(fields, 1):
            field_info = f"{i}. `{field['col']}列` **{field['name']}**"
            
            # 检查是否为筛选项
            if field['unique_values']:
                field_info += f" (筛选项: {', '.join(map(str, field['unique_values']))})"
            elif field['sample_values']:
                field_info += f" (示例: {', '.join(map(str, field['sample_values'][:3]))}...)"
            
            result.append(field_info)
        
        return result
    
    def _analyze_complex_table(self, ws, merged_ranges) -> list:
        """分析复杂表格"""
        result = []
        
        # 分析合并单元格结构
        merge_analysis = self._analyze_merged_structure(ws, merged_ranges)
        
        # 添加合并单元格详细信息
        if merged_ranges:
            result.append("**🔗 合并单元格信息**:")
            for i, merged_range in enumerate(merged_ranges[:15], 1):  # 最多显示15个
                min_row, min_col = merged_range.min_row, merged_range.min_col
                max_row, max_col = merged_range.max_row, merged_range.max_col
                
                start_cell = f"{get_column_letter(min_col)}{min_row}"
                end_cell = f"{get_column_letter(max_col)}{max_row}"
                
                cell = ws.cell(min_row, min_col)
                value = str(cell.value)[:30] if cell.value else "(空)"
                if len(str(cell.value or "")) > 30:
                    value += "..."
                
                result.append(f"  {i}. `{start_cell}:{end_cell}` → \"{value}\"")
            
            if len(merged_ranges) > 15:
                result.append(f"  ... 还有{len(merged_ranges) - 15}个合并单元格")
            result.append("")
        
        # 找到数据区域
        data_start = self._find_data_start(ws, merged_ranges)
        data_start_cell = f"{get_column_letter(data_start['col'])}{data_start['row']}"
        result.append(f"**📍 数据起始位置**: `{data_start_cell}` (第{data_start['row']}行第{data_start['col']}列)")
        
        # 提取复杂字段结构
        fields = self._extract_complex_fields(ws, merge_analysis, data_start)
        
        result.append("")
        result.append("**🏷️ 字段结构** (智能解析):")
        
        # 主要字段
        main_fields = [f for f in fields if f['importance'] == 'main']
        if main_fields:
            result.append("*核心字段* (排除合并单元格后):")
            for i, field in enumerate(main_fields, 1):
                # 添加字段起始位置信息
                field_start_pos = self._get_field_start_position(ws, field['col'], merged_ranges)
                field_info = f"  {i}. `{field['col']}列` **{field['name']}** _(从{field_start_pos}开始)_"
                
                if field['unique_values']:
                    field_info += f" (筛选项: {', '.join(map(str, field['unique_values']))})"
                elif field['sample_values']:
                    field_info += f" (示例: {', '.join(map(str, field['sample_values'][:2]))}...)"
                result.append(field_info)
        
        # 辅助字段
        aux_fields = [f for f in fields if f['importance'] == 'aux']
        if aux_fields:
            result.append("")
            result.append("*辅助字段*:")
            for field in aux_fields:
                result.append(f"  • `{field['col']}列` {field['name']}")
        
        # 字段关系说明
        if len(fields) > 8:
            result.append("")
            result.append(f"**📈 字段递进关系**: A列→B列→C列...→{fields[-1]['col']}列")
            result.append(f"F列之后的字段: {' → '.join([f['name'] for f in fields[5:]])}")
        
        return result
    
    def _analyze_merged_structure(self, ws, merged_ranges):
        """分析合并单元格结构"""
        merge_info = {}
        for merged_range in merged_ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            
            cell = ws.cell(min_row, min_col)
            value = str(cell.value) if cell.value else ""
            
            merge_info[f"{min_row}_{min_col}"] = {
                'range': (min_row, min_col, max_row, max_col),
                'value': value,
                'span_rows': max_row - min_row + 1,
                'span_cols': max_col - min_col + 1
            }
        
        return merge_info
    
    def _find_data_start(self, ws, merged_ranges):
        """找到真实数据开始的位置"""
        # 跳过标题和表头区域，找到第一个有连续数据的行
        for row in range(1, min(10, ws.max_row + 1)):
            consecutive_data = 0
            start_col = 1
            
            for col in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    consecutive_data += 1
                    if consecutive_data >= 5:  # 连续5个有数据的单元格
                        return {'row': row, 'col': start_col}
                else:
                    consecutive_data = 0
                    start_col = col + 1
        
        return {'row': 2, 'col': 1}  # 默认值
    
    def _extract_complex_fields(self, ws, merge_analysis, data_start):
        """提取复杂表格的字段"""
        fields = []
        header_rows = [2, 3]  # demo2中主要是第2、3行作为表头
        
        for col in range(1, min(ws.max_column + 1, 50)):  # 限制在50列内
            field_name = self._get_complex_field_name(ws, col, header_rows, merge_analysis)
            
            if field_name:
                # 判断字段重要性
                importance = 'main' if col <= 16 else 'aux'  # 前16列为主要字段
                
                # 提取样本数据
                sample_values, unique_values = self._extract_field_data(ws, col, data_start['row'])
                
                fields.append({
                    'col': get_column_letter(col),
                    'name': field_name,
                    'importance': importance,
                    'sample_values': sample_values,
                    'unique_values': unique_values
                })
        
        return fields
    
    def _get_complex_field_name(self, ws, col, header_rows, merge_analysis):
        """获取复杂表格的字段名"""
        names = []
        
        for row in header_rows:
            cell = ws.cell(row, col)
            if cell.value is not None:
                value = str(cell.value).strip()
                if value and len(value) > 0:
                    names.append(value)
        
        if not names:
            return None
        
        # 智能合并多个表头层级
        if len(names) == 1:
            return names[0]
        elif len(names) == 2:
            # 如果第二个名称是第一个的细分，用"-"连接
            if names[0] and names[1]:
                return f"{names[0]}-{names[1]}"
            else:
                return names[0] or names[1]
        else:
            return " | ".join([n for n in names if n])
    
    def _extract_standard_fields(self, ws, header_row):
        """提取标准表格字段"""
        fields = []
        
        for col in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(header_row, col)
            if cell.value is not None:
                field_name = str(cell.value).strip()
                if field_name:
                    # 提取字段数据（从表头下一行开始）
                    sample_values, unique_values = self._extract_field_data(ws, col, header_row + 1)
                    
                    fields.append({
                        'col': get_column_letter(col),
                        'name': field_name,
                        'sample_values': sample_values,
                        'unique_values': unique_values
                    })
        
        return fields
    
    def _extract_field_data(self, ws, col, start_row):
        """提取字段的样本数据和唯一值"""
        values = []
        
        # 提取前100行的数据作为样本
        for row in range(start_row, min(start_row + 100, ws.max_row + 1)):
            cell = ws.cell(row, col)
            if cell.value is not None:
                value = str(cell.value).strip()
                if value:
                    values.append(value)
        
        if not values:
            return [], []
        
        # 统计唯一值
        unique_counts = Counter(values)
        unique_values = list(unique_counts.keys())
        
        # 如果唯一值少于等于10个，返回所有唯一值作为筛选项
        if len(unique_values) <= 10:
            return [], unique_values[:10]  # 返回前10个作为筛选项
        else:
            return values[:5], []  # 返回前5个作为样本
    
    def _find_best_header_row(self, ws) -> int:
        """找到最佳表头行（基于数据一致性分析）"""
        # 优先使用数据一致性分析
        header_row = self._find_header_by_data_consistency(ws)
        if header_row:
            return header_row
        
        # 如果没找到，使用传统方法作为兜底
        best_row = 1
        best_score = 0
        
        for row in range(1, min(6, ws.max_row + 1)):
            score = 0
            non_empty = 0
            
            for col in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row, col)
                if cell.value is not None:
                    non_empty += 1
                    value_str = str(cell.value).strip()
                    
                    # 使用改进的表头检测
                    if self._looks_like_header(value_str):
                        score += 2.0
                    elif not self._is_numeric(value_str) and len(value_str) <= 20:
                        score += 1.0
                    else:
                        score += 0.3
            
            if non_empty > 0:
                final_score = score * non_empty
                if final_score > best_score:
                    best_score = final_score
                    best_row = row
        
        return best_row
    
    def _looks_like_header(self, value_str: str) -> bool:
        """判断字符串是否像表头字段名（通用方法）"""
        if not value_str or len(value_str) > 100:  # 太长的不太像字段名
            return False
        
        value_str = value_str.strip()
        
        # 1. 日期格式通常不是表头字段名
        if self._looks_like_date(value_str):
            return False
        
        # 2. 纯数字通常不是表头（除非很短）
        if self._is_numeric(value_str):
            if len(value_str) > 4:  # 超过4位数字不太像表头
                return False
            elif len(value_str) > 2 and not any(c.isalpha() for c in value_str):
                return False  # 纯数字且超过2位
        
        # 3. 包含复杂标点的不是表头
        complex_patterns = [
            '。', '，', '；', '：', '？', '！',  # 中文标点
            '.', ',', ';', '?', '!',           # 英文标点  
            '\n', '\r', '\t',                 # 换行符
            '-', '/', '\\', '|'               # 分隔符（日期时间常用）
        ]
        if any(pattern in value_str for pattern in complex_patterns):
            return False
        
        # 4. 排除明显的数据内容模式
        if self._looks_like_data_content(value_str):
            return False
        
        # 5. 长度特征：表头字段通常比较简洁
        if len(value_str) <= 20:
            return True
            
        return False
    
    def _looks_like_data_content(self, value_str: str) -> bool:
        """检查是否像数据内容而不是表头（通用方法，不依赖具体词汇）"""
        # 这个方法现在只做最基本的模式识别，避免硬编码
        return False  # 暂时禁用，让其他逻辑来判断
    
    def _looks_like_date(self, value_str: str) -> bool:
        """检查是否像日期格式"""
        import re
        date_patterns = [
            r'^\d{8}$',           # 20250527
            r'^\d{4}-\d{2}-\d{2}$', # 2025-05-27
            r'^\d{4}/\d{2}/\d{2}$', # 2025/05/27
            r'^\d{4}\.\d{2}\.\d{2}$' # 2025.05.27
        ]
        for pattern in date_patterns:
            if re.match(pattern, value_str):
                return True
        return False
    
    def _is_numeric(self, value_str: str) -> bool:
        """检查字符串是否为数字"""
        try:
            float(value_str)
            return True
        except (ValueError, TypeError):
            return False
    
    def _get_field_start_position(self, ws, col_letter, merged_ranges):
        """获取字段实际开始的位置（排除合并单元格）"""
        from openpyxl.utils import column_index_from_string
        col_num = column_index_from_string(col_letter)
        
        # 检查前几行，找到第一个非合并且有内容的单元格
        for row in range(1, min(10, ws.max_row + 1)):
            cell = ws.cell(row, col_num)
            
            # 检查是否在合并单元格中
            is_in_merge = False
            for merged_range in merged_ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and 
                    merged_range.min_col <= col_num <= merged_range.max_col):
                    is_in_merge = True
                    break
            
            # 如果不在合并单元格中且有内容，这就是起始位置
            if not is_in_merge and cell.value is not None:
                return f"{col_letter}{row}"
        
        # 如果没找到，返回默认位置
        return f"{col_letter}1"

# 使用示例
def main():
    """主函数 - 演示用法"""
    analyzer = AITabAnalyzer()
    
    print("🤖 AI分析Tab - 智能Excel分析器")
    print("=" * 60)
    
    # 分析demo1
    print("\n📄 demo1.xlsx 分析结果:")
    result1 = analyzer.analyze_for_ai('demo1.xlsx')
    print(result1)
    
    print("\n" + "="*60)
    
    # 分析demo2  
    print("\n📄 demo2.xlsx 分析结果:")
    result2 = analyzer.analyze_for_ai('demo2.xlsx')
    print(result2)

if __name__ == "__main__":
    main() 