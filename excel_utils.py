import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import numpy as np
from typing import Dict, List, Any, Tuple, Optional
import io
import tempfile
import os

class SmartExcelAnalyzer:
    """智能Excel分析器 - 自动识别和处理各种Excel文件结构"""
    
    def __init__(self):
        self.workbook = None
        self.analysis_result = {}
    
    def analyze_excel_structure(self, file_path: str) -> Dict[str, Any]:
        """全面分析Excel文件结构"""
        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            analysis = {
                'file_path': file_path,
                'total_sheets': len(self.workbook.sheetnames),
                'sheet_names': self.workbook.sheetnames,
                'sheets_analysis': {},
                'cross_sheet_analysis': {},
                'ai_prompt_data': {}
            }
            
            for sheet_name in self.workbook.sheetnames:
                sheet_analysis = self._analyze_worksheet(sheet_name)
                analysis['sheets_analysis'][sheet_name] = sheet_analysis
            
            # 跨工作表关联分析
            analysis['cross_sheet_analysis'] = self._analyze_cross_sheet_relationships(analysis['sheets_analysis'])
            
            # 生成AI提示词数据
            analysis['ai_prompt_data'] = self._generate_ai_prompt_data(analysis)
            
            self.analysis_result = analysis
            return analysis
            
        except Exception as e:
            raise Exception(f"分析Excel文件结构时出错: {str(e)}")
    
    def _analyze_worksheet(self, sheet_name: str) -> Dict[str, Any]:
        """分析单个工作表的结构"""
        ws = self.workbook[sheet_name]
        
        # 基本信息分析
        basic_info = {
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'total_cells': ws.max_row * ws.max_column
        }
        
        # 合并单元格分析
        merged_analysis = self._analyze_merged_cells(ws)
        
        # 数据区域检测
        data_regions = self._detect_data_regions(ws)
        
        # 表格结构识别
        table_structures = self._identify_table_structures(ws, data_regions)
        
        # 字段分析（核心新增功能）
        field_analysis = self._analyze_fields_and_content(ws, table_structures)
        
        # 生成读取建议
        read_suggestions = self._generate_read_suggestions(table_structures, merged_analysis)
        
        return {
            'basic_info': basic_info,
            'merged_cells': merged_analysis,
            'data_regions': data_regions,
            'table_structures': table_structures,
            'field_analysis': field_analysis,  # 新增
            'read_suggestions': read_suggestions
        }
    
    def _analyze_fields_and_content(self, ws, table_structures: List[Dict[str, Any]]) -> Dict[str, Any]:
        """分析字段名和数据内容 - 核心功能"""
        if not table_structures:
            return {"status": "no_table_detected", "fields": [], "data_samples": []}
        
        main_table = table_structures[0]
        region = main_table['region']
        header_analysis = main_table['header_analysis']
        
        min_row, max_row = region['min_row'], region['max_row']
        min_col, max_col = region['min_col'], region['max_col']
        header_row = header_analysis['suggested_header_row']
        
        # 提取字段信息
        fields_info = []
        
        for col in range(min_col, max_col + 1):
            field_info = self._analyze_single_field(ws, col, header_row, max_row)
            if field_info:
                fields_info.append(field_info)
        
        # 提取数据样本
        data_samples = self._extract_data_samples(ws, header_row + 1, max_row, min_col, max_col, limit=10)
        
        # 数据概述
        data_summary = self._generate_data_summary(fields_info, data_samples)
        
        return {
            "status": "success",
            "header_row": header_row,
            "fields_count": len(fields_info),
            "data_rows_count": max_row - header_row,
            "fields": fields_info,
            "data_samples": data_samples,
            "data_summary": data_summary
        }
    
    def _analyze_single_field(self, ws, col: int, header_row: int, max_row: int) -> Dict[str, Any]:
        """分析单个字段的详细信息"""
        # 获取字段名
        header_cell = ws.cell(row=header_row, column=col)
        field_name = str(header_cell.value).strip() if header_cell.value else f"字段{col}"
        
        # 收集该列的数据样本（最多50个）
        data_values = []
        sample_limit = min(50, max_row - header_row)
        
        for row in range(header_row + 1, header_row + 1 + sample_limit):
            if row > max_row:
                break
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                data_values.append(cell.value)
        
        # 分析数据类型和特征
        field_analysis = self._deep_analyze_field_data(data_values)
        
        return {
            "column_index": col,
            "column_letter": get_column_letter(col),
            "field_name": field_name,
            "original_field_name": str(header_cell.value) if header_cell.value else "",
            "data_count": len(data_values),
            "sample_values": [str(v)[:50] for v in data_values[:5]],  # 前5个样本
            "data_type_analysis": field_analysis,
            "is_key_field": self._is_potential_key_field(field_name, data_values),
            "field_patterns": self._detect_field_patterns(data_values)
        }
    
    def _deep_analyze_field_data(self, values: List[Any]) -> Dict[str, Any]:
        """深度分析字段数据类型和特征"""
        if not values:
            return {"primary_type": "empty", "confidence": 1.0, "characteristics": []}
        
        analysis = {
            "total_values": len(values),
            "unique_values": len(set(str(v) for v in values)),
            "null_count": 0,
            "type_distribution": {"numeric": 0, "date": 0, "text": 0, "boolean": 0, "id": 0},
            "characteristics": [],
            "value_range": {},
            "common_values": []
        }
        
        # 类型检测
        for value in values:
            if value is None or str(value).strip() == "":
                analysis["null_count"] += 1
                continue
                
            str_value = str(value).strip()
            
            # ID类型检测
            if self._looks_like_id(str_value):
                analysis["type_distribution"]["id"] += 1
            # 数值检测
            elif self._is_numeric(value):
                analysis["type_distribution"]["numeric"] += 1
            # 日期检测
            elif self._is_date_like(str_value):
                analysis["type_distribution"]["date"] += 1
            # 布尔值检测
            elif self._is_boolean_like(str_value):
                analysis["type_distribution"]["boolean"] += 1
            else:
                analysis["type_distribution"]["text"] += 1
        
        # 确定主要类型
        valid_count = len(values) - analysis["null_count"]
        if valid_count > 0:
            primary_type = max(analysis["type_distribution"].keys(), 
                             key=lambda k: analysis["type_distribution"][k])
            confidence = analysis["type_distribution"][primary_type] / valid_count
            
            # 数值分析
            if primary_type == "numeric":
                numeric_values = [float(v) for v in values if self._is_numeric(v)]
                if numeric_values:
                    analysis["value_range"] = {
                        "min": min(numeric_values),
                        "max": max(numeric_values),
                        "avg": sum(numeric_values) / len(numeric_values)
                    }
            
            # 文本长度分析
            if primary_type == "text":
                text_lengths = [len(str(v)) for v in values if v is not None]
                if text_lengths:
                    analysis["value_range"] = {
                        "min_length": min(text_lengths),
                        "max_length": max(text_lengths),
                        "avg_length": sum(text_lengths) / len(text_lengths)
                    }
            
            # 常见值统计
            from collections import Counter
            value_counts = Counter(str(v) for v in values if v is not None)
            analysis["common_values"] = value_counts.most_common(3)
            
            # 特征分析
            if analysis["unique_values"] == analysis["total_values"] - analysis["null_count"]:
                analysis["characteristics"].append("唯一值字段")
            
            if primary_type == "id":
                analysis["characteristics"].append("疑似ID字段")
            
            if analysis["null_count"] > len(values) * 0.3:
                analysis["characteristics"].append("存在大量空值")
                
        else:
            primary_type = "empty"
            confidence = 1.0
        
        analysis["primary_type"] = primary_type
        analysis["confidence"] = confidence
        
        return analysis
    
    def _analyze_merged_cells(self, ws) -> Dict[str, Any]:
        """分析合并单元格模式"""
        merged_ranges = list(ws.merged_cells.ranges)
        
        analysis = {
            'count': len(merged_ranges),
            'header_merges': 0,
            'data_merges': 0,
            'complex_structure': False,
            'ranges_info': []
        }
        
        for merged_range in merged_ranges:
            range_info = {
                'range': str(merged_range),
                'start_row': merged_range.min_row,
                'end_row': merged_range.max_row,
                'start_col': merged_range.min_col,
                'end_col': merged_range.max_col,
                'span_rows': merged_range.max_row - merged_range.min_row + 1,
                'span_cols': merged_range.max_col - merged_range.min_col + 1
            }
            
            # 获取合并单元格的值
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            range_info['value'] = str(top_left_cell.value)[:50] if top_left_cell.value else ""
            
            # 判断是标题合并还是数据合并
            if merged_range.min_row <= 5:
                analysis['header_merges'] += 1
                range_info['type'] = 'header'
            else:
                analysis['data_merges'] += 1
                range_info['type'] = 'data'
            
            analysis['ranges_info'].append(range_info)
        
        analysis['complex_structure'] = len(merged_ranges) > 10 or analysis['header_merges'] > 5
        
        return analysis
    
    def _detect_data_regions(self, ws) -> List[Dict[str, Any]]:
        """智能检测数据区域"""
        data_cells = []
        
        # 扫描工作表，限制扫描范围以提高性能
        max_scan_row = min(ws.max_row, 500)
        max_scan_col = min(ws.max_column, 100)
        
        for row in range(1, max_scan_row + 1):
            for col in range(1, max_scan_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    data_cells.append((row, col, str(cell.value).strip()))
        
        if not data_cells:
            return []
        
        # 分析数据分布模式
        return self._cluster_data_regions(data_cells)
    
    def _cluster_data_regions(self, data_cells: List[Tuple[int, int, str]]) -> List[Dict[str, Any]]:
        """聚类数据单元格形成有意义的区域"""
        if not data_cells:
            return []
        
        sorted_cells = sorted(data_cells, key=lambda x: (x[0], x[1]))
        
        # 计算整体数据边界
        min_row = min(cell[0] for cell in sorted_cells)
        max_row = max(cell[0] for cell in sorted_cells)
        min_col = min(cell[1] for cell in sorted_cells)
        max_col = max(cell[1] for cell in sorted_cells)
        
        # 计算数据密度
        total_area = (max_row - min_row + 1) * (max_col - min_col + 1)
        data_density = len(data_cells) / total_area if total_area > 0 else 0
        
        # 分析行列数据分布
        row_data_count = {}
        col_data_count = {}
        
        for row, col, value in data_cells:
            row_data_count[row] = row_data_count.get(row, 0) + 1
            col_data_count[col] = col_data_count.get(col, 0) + 1
        
        # 识别主要数据区域
        main_region = {
            'id': 1,
            'min_row': min_row,
            'max_row': max_row,
            'min_col': min_col,
            'max_col': max_col,
            'cell_count': len(data_cells),
            'density': data_density,
            'row_distribution': row_data_count,
            'col_distribution': col_data_count,
            'type': self._classify_region_type(data_density, row_data_count, col_data_count),
            'sample_values': [cell[2][:30] for cell in sorted_cells[:5]]
        }
        
        return [main_region]
    
    def _classify_region_type(self, density: float, row_dist: Dict[int, int], col_dist: Dict[int, int]) -> str:
        """分类数据区域类型"""
        if density > 0.7:
            return "dense_table"
        elif density > 0.3:
            return "standard_table"
        else:
            return "sparse_data"
    
    def _identify_table_structures(self, ws, data_regions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """识别表格结构"""
        table_structures = []
        
        for region in data_regions:
            table_info = self._analyze_table_structure(ws, region)
            if table_info:
                table_structures.append(table_info)
        
        return table_structures
    
    def _analyze_table_structure(self, ws, region: Dict[str, Any]) -> Dict[str, Any]:
        """深度分析表格结构"""
        min_row, max_row = region['min_row'], region['max_row']
        min_col, max_col = region['min_col'], region['max_col']
        
        # 智能识别表头位置
        header_analysis = self._smart_header_detection(ws, min_row, max_row, min_col, max_col)
        
        # 检测数据开始行
        data_start_row = self._find_data_start_row(ws, min_row, max_row, min_col, max_col, header_analysis)
        
        # 分析表格类型和方向
        table_type = self._classify_table_type(max_row - min_row + 1, max_col - min_col + 1)
        
        # 列数据类型分析
        column_analysis = self._analyze_column_types(ws, data_start_row, max_row, min_col, max_col)
        
        return {
            'region': region,
            'header_analysis': header_analysis,
            'data_start_row': data_start_row,
            'table_type': table_type,
            'column_analysis': column_analysis,
            'data_quality': self._assess_data_quality(ws, data_start_row, max_row, min_col, max_col)
        }
    
    def _smart_header_detection(self, ws, min_row: int, max_row: int, min_col: int, max_col: int) -> Dict[str, Any]:
        """智能检测表头"""
        header_candidates = []
        
        # 检查前10行的内容特征
        for row in range(min_row, min(min_row + 10, max_row + 1)):
            row_analysis = {
                'row': row,
                'text_ratio': 0,
                'unique_values': 0,
                'formatting_score': 0,
                'content_sample': []
            }
            
            values_in_row = []
            text_count = 0
            
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    value = str(cell.value).strip()
                    values_in_row.append(value)
                    
                    # 检查是否为文本（非纯数字）
                    try:
                        float(value)
                    except (ValueError, TypeError):
                        text_count += 1
                    
                    # 检查格式特征（粗体、颜色等）
                    if hasattr(cell, 'font') and cell.font and cell.font.bold:
                        row_analysis['formatting_score'] += 1
            
            if values_in_row:
                row_analysis['text_ratio'] = text_count / len(values_in_row)
                row_analysis['unique_values'] = len(set(values_in_row))
                row_analysis['content_sample'] = values_in_row[:5]
                
                # 判断是否可能是表头
                if (row_analysis['text_ratio'] > 0.6 or  # 大部分是文本
                    row_analysis['formatting_score'] > 0 or  # 有格式特征
                    len(set(values_in_row)) == len(values_in_row)):  # 全部都是唯一值
                    header_candidates.append(row_analysis)
        
        # 选择最佳表头行
        best_header_row = min_row
        if header_candidates:
            # 综合评分选择最佳表头
            best_candidate = max(header_candidates, 
                               key=lambda x: x['text_ratio'] * 0.4 + 
                                           (x['formatting_score'] > 0) * 0.3 + 
                                           (x['unique_values'] > 1) * 0.3)
            best_header_row = best_candidate['row']
        
        return {
            'candidates': header_candidates,
            'suggested_header_row': best_header_row,
            'confidence': len(header_candidates) / max(1, min(10, max_row - min_row + 1))
        }
    
    def _find_data_start_row(self, ws, min_row: int, max_row: int, min_col: int, max_col: int, header_analysis: Dict) -> int:
        """找到数据开始行"""
        suggested_header = header_analysis.get('suggested_header_row', min_row)
        
        # 从建议的表头行后开始查找数据
        for row in range(suggested_header + 1, max_row + 1):
            consecutive_data = 0
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    consecutive_data += 1
                else:
                    break
            
            # 如果该行有连续的数据，认为是数据开始行
            if consecutive_data >= (max_col - min_col + 1) * 0.5:
                return row
        
        return suggested_header + 1 if suggested_header < max_row else min_row
    
    def _classify_table_type(self, rows: int, cols: int) -> str:
        """分类表格类型"""
        if rows > cols * 4:
            return "vertical_long"  # 纵向长表
        elif cols > rows * 4:
            return "horizontal_wide"  # 横向宽表
        elif rows > 50 and cols > 20:
            return "large_matrix"  # 大型矩阵表
        elif rows <= 10 or cols <= 3:
            return "small_table"  # 小表格
        else:
            return "standard_table"  # 标准表格
    
    def _analyze_column_types(self, ws, start_row: int, end_row: int, start_col: int, end_col: int) -> List[Dict[str, Any]]:
        """分析每列的数据类型"""
        columns_info = []
        
        for col in range(start_col, end_col + 1):
            values = []
            # 采样分析（最多取50个值）
            sample_rows = list(range(start_row, min(start_row + 50, end_row + 1)))
            
            for row in sample_rows:
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    values.append(cell.value)
            
            if values:
                col_info = {
                    'column_index': col,
                    'column_letter': get_column_letter(col),
                    'sample_size': len(values),
                    'data_type': self._detect_column_data_type(values),
                    'null_ratio': (len(sample_rows) - len(values)) / len(sample_rows),
                    'unique_ratio': len(set(values)) / len(values),
                    'sample_values': [str(v)[:20] for v in values[:3]]
                }
                columns_info.append(col_info)
        
        return columns_info
    
    def _detect_column_data_type(self, values: List[Any]) -> Dict[str, Any]:
        """智能检测列数据类型"""
        if not values:
            return {"primary_type": "empty", "confidence": 1.0}
        
        type_counts = {"numeric": 0, "date": 0, "text": 0, "boolean": 0, "mixed": 0}
        
        for value in values[:20]:  # 检查前20个值
            str_value = str(value).strip().lower()
            
            # 数值检测
            try:
                float(value)
                type_counts["numeric"] += 1
                continue
            except (ValueError, TypeError):
                pass
            
            # 布尔值检测
            if str_value in ['true', 'false', '是', '否', 'yes', 'no', '1', '0']:
                type_counts["boolean"] += 1
                continue
            
            # 日期检测
            try:
                pd.to_datetime(str_value)
                type_counts["date"] += 1
                continue
            except:
                pass
            
            # 默认为文本
            type_counts["text"] += 1
        
        # 确定主要类型
        total_checked = sum(type_counts.values())
        primary_type = max(type_counts.keys(), key=lambda k: type_counts[k])
        confidence = type_counts[primary_type] / total_checked if total_checked > 0 else 0
        
        return {
            "primary_type": primary_type,
            "confidence": confidence,
            "type_distribution": type_counts
        }
    
    def _assess_data_quality(self, ws, start_row: int, end_row: int, start_col: int, end_col: int) -> Dict[str, Any]:
        """评估数据质量"""
        total_cells = (end_row - start_row + 1) * (end_col - start_col + 1)
        empty_cells = 0
        
        # 采样检查数据质量
        sample_size = min(1000, total_cells)
        checked_cells = 0
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                if checked_cells >= sample_size:
                    break
                
                cell = ws.cell(row=row, column=col)
                if cell.value is None or str(cell.value).strip() == "":
                    empty_cells += 1
                
                checked_cells += 1
            
            if checked_cells >= sample_size:
                break
        
        completeness = 1 - (empty_cells / checked_cells) if checked_cells > 0 else 0
        
        return {
            "completeness": completeness,
            "total_cells_sampled": checked_cells,
            "empty_cells": empty_cells,
            "quality_score": completeness  # 可以扩展更多质量指标
        }
    
    def _generate_read_suggestions(self, table_structures: List[Dict[str, Any]], merged_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """生成智能读取建议"""
        if not table_structures:
            return {
                "method": "simple_read",
                "parameters": {},
                "warnings": ["无法检测到有效的表格结构"]
            }
        
        main_table = table_structures[0]
        header_analysis = main_table.get('header_analysis', {})
        
        suggestions = {
            "method": "intelligent_read",
            "parameters": {
                "header": header_analysis.get('suggested_header_row', 1) - 1,  # pandas使用0-based
                "skiprows": header_analysis.get('suggested_header_row', 1) - 1,
                "nrows": None,
                "usecols": None
            },
            "preprocessing_steps": [],
            "warnings": [],
            "confidence": header_analysis.get('confidence', 0.5)
        }
        
        # 基于表格类型添加建议
        table_type = main_table.get('table_type', 'standard_table')
        if table_type == "horizontal_wide":
            suggestions["preprocessing_steps"].append("考虑转置表格")
            suggestions["warnings"].append("检测到宽表格，可能需要转置处理")
        elif table_type == "large_matrix":
            suggestions["preprocessing_steps"].append("分块读取大型数据")
            suggestions["warnings"].append("检测到大型表格，建议分批处理")
        
        # 基于合并单元格情况添加建议
        if merged_analysis.get('complex_structure', False):
            suggestions["preprocessing_steps"].append("处理复杂的合并单元格")
            suggestions["warnings"].append("存在复杂的合并单元格结构，可能影响数据读取")
        
        # 基于数据质量添加建议
        data_quality = main_table.get('data_quality', {})
        if data_quality.get('completeness', 1) < 0.8:
            suggestions["preprocessing_steps"].append("处理缺失值")
            suggestions["warnings"].append(f"数据完整性较低({data_quality.get('completeness', 0):.1%})，存在较多空值")
        
        return suggestions
    
    def _is_numeric(self, value) -> bool:
        """检测是否为数值"""
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False
    
    def _is_date_like(self, value: str) -> bool:
        """检测是否像日期"""
        try:
            pd.to_datetime(value)
            return True
        except:
            # 检查常见日期格式
            import re
            date_patterns = [
                r'\d{4}-\d{1,2}-\d{1,2}',
                r'\d{4}/\d{1,2}/\d{1,2}',
                r'\d{1,2}-\d{1,2}-\d{4}',
                r'\d{1,2}/\d{1,2}/\d{4}'
            ]
            return any(re.match(pattern, value) for pattern in date_patterns)
    
    def _is_boolean_like(self, value: str) -> bool:
        """检测是否像布尔值"""
        return value.lower() in ['true', 'false', '是', '否', 'yes', 'no', '1', '0', '✓', '×']
    
    def _looks_like_id(self, value: str) -> bool:
        """检测是否像ID"""
        import re
        # ID特征：连续数字、包含字母数字组合、特定格式
        patterns = [
            r'^\d+$',  # 纯数字
            r'^[A-Za-z]\d+$',  # 字母+数字
            r'^[A-Za-z]{2,}\d+$',  # 多字母+数字
            r'^\d+-\d+$',  # 数字-数字
            r'^[A-Za-z0-9]{8,}$'  # 长字符串
        ]
        return any(re.match(pattern, value) for pattern in patterns)
    
    def _is_potential_key_field(self, field_name: str, values: List[Any]) -> bool:
        """判断是否为潜在的关键字段"""
        field_name_lower = field_name.lower()
        key_indicators = ['id', 'key', 'code', '编号', '代码', '序号', 'uuid', 'guid']
        
        # 字段名包含关键指示词
        name_indicates_key = any(indicator in field_name_lower for indicator in key_indicators)
        
        # 数据特征分析
        unique_ratio = len(set(str(v) for v in values)) / len(values) if values else 0
        is_mostly_unique = unique_ratio > 0.8
        
        return name_indicates_key or is_mostly_unique
    
    def _detect_field_patterns(self, values: List[Any]) -> List[str]:
        """检测字段的模式特征"""
        patterns = []
        
        if not values:
            return patterns
        
        # 检查常见模式
        str_values = [str(v) for v in values if v is not None]
        
        if str_values:
            # 长度一致性
            lengths = [len(s) for s in str_values]
            if len(set(lengths)) == 1:
                patterns.append(f"固定长度({lengths[0]}字符)")
            
            # 格式模式
            import re
            if all(re.match(r'^\d{4}-\d{2}-\d{2}$', s) for s in str_values[:5]):
                patterns.append("日期格式(YYYY-MM-DD)")
            elif all(re.match(r'^\d+$', s) for s in str_values[:5]):
                patterns.append("纯数字")
            elif all(re.match(r'^[A-Za-z]+\d+$', s) for s in str_values[:5]):
                patterns.append("字母+数字组合")
        
        return patterns
    
    def _extract_data_samples(self, ws, start_row: int, end_row: int, start_col: int, end_col: int, limit: int = 10) -> List[Dict[str, Any]]:
        """提取数据样本"""
        samples = []
        
        # 获取表头
        headers = []
        for col in range(start_col, end_col + 1):
            header_cell = ws.cell(row=start_row - 1, column=col)
            header = str(header_cell.value).strip() if header_cell.value else f"列{col}"
            headers.append(header)
        
        # 提取数据行
        sample_count = 0
        for row in range(start_row, end_row + 1):
            if sample_count >= limit:
                break
            
            row_data = {}
            has_data = False
            
            for i, col in enumerate(range(start_col, end_col + 1)):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                
                if value is not None:
                    has_data = True
                    # 限制显示长度
                    if isinstance(value, str) and len(value) > 100:
                        value = value[:100] + "..."
                
                row_data[headers[i]] = value
            
            if has_data:
                row_data["_row_number"] = row
                samples.append(row_data)
                sample_count += 1
        
        return samples
    
    def _generate_data_summary(self, fields_info: List[Dict], data_samples: List[Dict]) -> Dict[str, Any]:
        """生成数据概述"""
        summary = {
            "total_fields": len(fields_info),
            "key_fields": [],
            "numeric_fields": [],
            "text_fields": [],
            "date_fields": [],
            "id_fields": [],
            "fields_with_issues": [],
            "sample_row_count": len(data_samples)
        }
        
        for field in fields_info:
            field_name = field["field_name"]
            data_type = field["data_type_analysis"]["primary_type"]
            characteristics = field["data_type_analysis"]["characteristics"]
            
            # 分类字段
            if field["is_key_field"]:
                summary["key_fields"].append(field_name)
            
            if data_type == "numeric":
                summary["numeric_fields"].append(field_name)
            elif data_type == "text":
                summary["text_fields"].append(field_name)
            elif data_type == "date":
                summary["date_fields"].append(field_name)
            elif data_type == "id":
                summary["id_fields"].append(field_name)
            
            # 问题字段
            if "存在大量空值" in characteristics:
                summary["fields_with_issues"].append({
                    "field": field_name,
                    "issue": "存在大量空值",
                    "null_ratio": field["data_type_analysis"]["null_count"] / field["data_type_analysis"]["total_values"]
                })
        
        return summary
    
    def _analyze_cross_sheet_relationships(self, sheets_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """分析跨工作表关联关系"""
        relationships = {
            "potential_links": [],
            "common_fields": [],
            "sheet_dependencies": [],
            "data_flow_analysis": {}
        }
        
        # 收集所有字段信息
        all_fields = {}
        for sheet_name, analysis in sheets_analysis.items():
            field_analysis = analysis.get('field_analysis', {})
            if field_analysis.get('status') == 'success':
                fields = field_analysis.get('fields', [])
                all_fields[sheet_name] = [f["field_name"] for f in fields]
        
        # 寻找相似字段名
        for sheet1, fields1 in all_fields.items():
            for sheet2, fields2 in all_fields.items():
                if sheet1 >= sheet2:  # 避免重复比较
                    continue
                
                common = set(fields1) & set(fields2)
                if common:
                    relationships["common_fields"].append({
                        "sheet1": sheet1,
                        "sheet2": sheet2,
                        "common_fields": list(common),
                        "similarity_score": len(common) / max(len(fields1), len(fields2))
                    })
        
        # 分析潜在的主从关系
        relationships["data_flow_analysis"] = self._analyze_data_flow_patterns(sheets_analysis)
        
        return relationships
    
    def _analyze_data_flow_patterns(self, sheets_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """分析数据流模式"""
        patterns = {
            "source_sheets": [],  # 源数据表
            "summary_sheets": [],  # 汇总表
            "detail_sheets": [],  # 明细表
            "template_sheets": []  # 模板表
        }
        
        for sheet_name, analysis in sheets_analysis.items():
            basic_info = analysis.get('basic_info', {})
            field_analysis = analysis.get('field_analysis', {})
            
            row_count = basic_info.get('max_row', 0)
            col_count = basic_info.get('max_column', 0)
            
            # 基于数据量和字段特征判断表类型
            if row_count > 1000:
                patterns["source_sheets"].append({
                    "sheet": sheet_name,
                    "reason": f"大量数据行({row_count}行)",
                    "rows": row_count,
                    "cols": col_count
                })
            elif "汇总" in sheet_name or "总计" in sheet_name or "统计" in sheet_name:
                patterns["summary_sheets"].append({
                    "sheet": sheet_name,
                    "reason": "表名包含汇总关键词",
                    "rows": row_count,
                    "cols": col_count
                })
            elif "模板" in sheet_name or "template" in sheet_name.lower():
                patterns["template_sheets"].append({
                    "sheet": sheet_name,
                    "reason": "表名包含模板关键词",
                    "rows": row_count,
                    "cols": col_count
                })
            elif row_count < 100 and col_count > 10:
                patterns["detail_sheets"].append({
                    "sheet": sheet_name,
                    "reason": "宽表格式，可能是明细或配置表",
                    "rows": row_count,
                    "cols": col_count
                })
        
        return patterns
    
    def _generate_ai_prompt_data(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """生成供AI分析的完整提示词数据"""
        ai_data = {
            "file_overview": {
                "file_path": analysis['file_path'],
                "total_sheets": analysis['total_sheets'],
                "sheet_names": analysis['sheet_names']
            },
            "detailed_sheet_analysis": {},
            "cross_sheet_relationships": analysis['cross_sheet_analysis'],
            "analysis_suggestions": [],
            "potential_analysis_tasks": [],
            "generated_prompt": ""
        }
        
        # 详细的工作表分析
        for sheet_name, sheet_analysis in analysis['sheets_analysis'].items():
            field_analysis = sheet_analysis.get('field_analysis', {})
            
            if field_analysis.get('status') == 'success':
                sheet_data = {
                    "sheet_name": sheet_name,
                    "dimensions": f"{sheet_analysis['basic_info']['max_row']}行 × {sheet_analysis['basic_info']['max_column']}列",
                    "field_count": field_analysis['fields_count'],
                    "data_rows": field_analysis['data_rows_count'],
                    "fields_detail": [],
                    "data_samples": field_analysis['data_samples'][:3],  # 前3行样本
                    "data_summary": field_analysis['data_summary'],
                    "special_notes": []
                }
                
                # 字段详情
                for field in field_analysis['fields']:
                    field_detail = {
                        "name": field['field_name'],
                        "type": field['data_type_analysis']['primary_type'],
                        "confidence": field['data_type_analysis']['confidence'],
                        "characteristics": field['data_type_analysis']['characteristics'],
                        "sample_values": field['sample_values'],
                        "is_key": field['is_key_field']
                    }
                    sheet_data["fields_detail"].append(field_detail)
                
                # 特殊注意事项
                merged_info = sheet_analysis.get('merged_cells', {})
                if merged_info.get('count', 0) > 0:
                    sheet_data["special_notes"].append(f"存在{merged_info['count']}个合并单元格")
                
                if merged_info.get('complex_structure', False):
                    sheet_data["special_notes"].append("复杂的合并单元格结构，可能影响数据读取")
                
                ai_data["detailed_sheet_analysis"][sheet_name] = sheet_data
        
        # 生成分析建议
        ai_data["analysis_suggestions"] = self._generate_analysis_suggestions(analysis)
        
        # 生成潜在分析任务
        ai_data["potential_analysis_tasks"] = self._generate_potential_tasks(analysis)
        
        # 生成结构化提示词
        ai_data["generated_prompt"] = self._generate_structured_prompt(ai_data)
        
        return ai_data
    
    def _generate_analysis_suggestions(self, analysis: Dict[str, Any]) -> List[str]:
        """生成分析建议"""
        suggestions = []
        
        # 基于跨表关系生成建议
        cross_analysis = analysis.get('cross_sheet_analysis', {})
        common_fields = cross_analysis.get('common_fields', [])
        
        if common_fields:
            suggestions.append("检测到多个工作表存在相同字段，建议分析表间关联关系")
        
        # 基于数据流分析生成建议
        data_flow = cross_analysis.get('data_flow_analysis', {})
        if data_flow.get('source_sheets'):
            suggestions.append("检测到源数据表，建议作为主要分析对象")
        
        if data_flow.get('summary_sheets'):
            suggestions.append("检测到汇总表，建议用于验证分析结果")
        
        return suggestions
    
    def _generate_potential_tasks(self, analysis: Dict[str, Any]) -> List[Dict[str, str]]:
        """生成潜在的分析任务"""
        tasks = []
        
        for sheet_name, sheet_analysis in analysis['sheets_analysis'].items():
            field_analysis = sheet_analysis.get('field_analysis', {})
            
            if field_analysis.get('status') == 'success':
                summary = field_analysis.get('data_summary', {})
                
                # 数据清洗任务
                if summary.get('fields_with_issues'):
                    tasks.append({
                        "type": "数据清洗",
                        "description": f"处理{sheet_name}中的数据质量问题",
                        "sheet": sheet_name
                    })
                
                # 统计分析任务
                if summary.get('numeric_fields'):
                    tasks.append({
                        "type": "统计分析",
                        "description": f"对{sheet_name}中的数值字段进行统计分析",
                        "sheet": sheet_name
                    })
                
                # 关联分析任务
                if summary.get('key_fields'):
                    tasks.append({
                        "type": "关联分析",
                        "description": f"基于{sheet_name}的关键字段进行跨表关联",
                        "sheet": sheet_name
                    })
        
        return tasks
    
    def _generate_structured_prompt(self, ai_data: Dict[str, Any]) -> str:
        """生成结构化的AI分析提示词"""
        prompt_parts = []
        
        # 文件概述
        overview = ai_data["file_overview"]
        prompt_parts.append(f"# Excel文件分析 - {overview['file_path']}")
        prompt_parts.append(f"文件包含{overview['total_sheets']}个工作表：{', '.join(overview['sheet_names'])}")
        prompt_parts.append("")
        
        # 详细工作表信息
        prompt_parts.append("## 工作表详细信息")
        
        for sheet_name, sheet_data in ai_data["detailed_sheet_analysis"].items():
            prompt_parts.append(f"### 【{sheet_name}】")
            prompt_parts.append(f"- **规模**: {sheet_data['dimensions']}")
            prompt_parts.append(f"- **字段数量**: {sheet_data['field_count']}")
            prompt_parts.append(f"- **数据行数**: {sheet_data['data_rows']}")
            
            # 字段信息
            prompt_parts.append("- **字段详情**:")
            for field in sheet_data["fields_detail"]:
                characteristics = ", ".join(field["characteristics"]) if field["characteristics"] else "无特殊特征"
                prompt_parts.append(f"  - `{field['name']}`: {field['type']} (置信度:{field['confidence']:.1%}) - {characteristics}")
                if field["sample_values"]:
                    sample_str = ", ".join([f'"{v}"' for v in field["sample_values"][:3]])
                    prompt_parts.append(f"    样本值: {sample_str}")
            
            # 数据样本
            if sheet_data["data_samples"]:
                prompt_parts.append("- **数据样本**:")
                for i, sample in enumerate(sheet_data["data_samples"][:2], 1):
                    sample_str = ", ".join([f"{k}={v}" for k, v in list(sample.items())[:5] if k != "_row_number"])
                    prompt_parts.append(f"  样本{i}: {sample_str}")
            
            # 特殊注意事项
            if sheet_data["special_notes"]:
                prompt_parts.append("- **注意事项**: " + "; ".join(sheet_data["special_notes"]))
            
            prompt_parts.append("")
        
        # 跨表关系
        relationships = ai_data["cross_sheet_relationships"]
        if relationships.get("common_fields"):
            prompt_parts.append("## 工作表关联关系")
            for rel in relationships["common_fields"]:
                prompt_parts.append(f"- {rel['sheet1']} ↔ {rel['sheet2']}: 共同字段 {rel['common_fields']}")
            prompt_parts.append("")
        
        # 分析建议
        if ai_data["analysis_suggestions"]:
            prompt_parts.append("## 分析建议")
            for suggestion in ai_data["analysis_suggestions"]:
                prompt_parts.append(f"- {suggestion}")
            prompt_parts.append("")
        
        # 潜在任务
        if ai_data["potential_analysis_tasks"]:
            prompt_parts.append("## 建议的分析任务")
            for task in ai_data["potential_analysis_tasks"]:
                prompt_parts.append(f"- **{task['type']}**: {task['description']}")
            prompt_parts.append("")
        
        return "\n".join(prompt_parts)

class AdvancedExcelProcessor:
    """增强版Excel处理类 - 整合智能分析功能"""
    
    def __init__(self):
        self.workbook = None
        self.file_path = None
        self.modified_data = {}
        self.analyzer = SmartExcelAnalyzer()
        self.structure_analysis = None
    
    def load_excel(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """加载Excel文件，保证数据完整性，不跳过任何行"""
        self.file_path = file_path
        
        # 首先进行结构分析（但不影响数据读取）
        print("🔍 正在分析Excel文件结构...")
        try:
            self.structure_analysis = self.analyzer.analyze_excel_structure(file_path)
        except Exception as e:
            print(f"⚠️ 结构分析失败，继续使用基础读取: {str(e)}")
            self.structure_analysis = None
        
        # 使用完整数据读取方法，确保不丢失任何数据
        excel_data = {}
        
        # 获取所有工作表名称
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        
        for sheet_name in sheet_names:
            print(f"📋 正在处理工作表: {sheet_name}")
            
            try:
                # 使用完整读取方法，确保从第一行开始读取所有数据
                df, markdown = self.read_excel_with_merged_cells_complete(file_path, sheet_name)
                excel_data[sheet_name] = df
                self.modified_data[sheet_name] = df.copy()
                
                # 打印读取摘要
                print(f"  ✅ 成功读取: {len(df)}行 × {len(df.columns)}列 (完整数据)")
                
            except Exception as e:
                print(f"  ❌ 读取出错: {str(e)}")
                # 最后的回退方法
                try:
                    df, markdown = self.read_excel_with_merged_cells(file_path, sheet_name)
                    excel_data[sheet_name] = df
                    self.modified_data[sheet_name] = df.copy()
                    print(f"  ✅ 回退方法成功: {len(df)}行 × {len(df.columns)}列")
                except Exception as e2:
                    print(f"  ❌ 回退方法也失败: {str(e2)}")
                    # 创建空DataFrame
                    excel_data[sheet_name] = pd.DataFrame()
                    self.modified_data[sheet_name] = pd.DataFrame()
        
        return excel_data
    
    def _smart_read_sheet(self, file_path: str, sheet_name: str, suggestions: Dict[str, Any]) -> Tuple[pd.DataFrame, str]:
        """基于智能分析建议读取工作表"""
        parameters = suggestions.get('parameters', {})
        
        # 构建pandas读取参数
        pd_params = {
            'io': file_path,
            'sheet_name': sheet_name,
            'engine': 'openpyxl'
        }
        
        # 应用智能建议的参数
        if parameters.get('header') is not None:
            pd_params['header'] = parameters['header']
        
        if parameters.get('skiprows') is not None and parameters['skiprows'] > 0:
            pd_params['skiprows'] = parameters['skiprows']
        
        if parameters.get('nrows') is not None:
            pd_params['nrows'] = parameters['nrows']
        
        if parameters.get('usecols') is not None:
            pd_params['usecols'] = parameters['usecols']
        
        # 读取数据
        df = pd.read_excel(**pd_params)
        
        # 如果需要，应用预处理步骤
        preprocessing_steps = suggestions.get('preprocessing_steps', [])
        for step in preprocessing_steps:
            if "转置" in step:
                df = df.transpose()
            elif "缺失值" in step:
                df = self._handle_missing_values(df)
        
        # 生成markdown预览
        markdown = self.df_to_markdown(df, sheet_name)
        
        return df, markdown
    
    def _handle_missing_values(self, df: pd.DataFrame) -> pd.DataFrame:
        """智能处理缺失值"""
        # 简单的缺失值处理策略
        # 数值列：用中位数填充
        # 文本列：用"未知"填充
        for col in df.columns:
            if df[col].dtype in ['int64', 'float64']:
                df[col].fillna(df[col].median(), inplace=True)
            else:
                df[col].fillna('未知', inplace=True)
        
        return df
    
    def get_structure_analysis(self) -> Dict[str, Any]:
        """获取结构分析结果"""
        return self.structure_analysis
    
    def print_analysis_summary(self):
        """打印分析摘要"""
        if not self.structure_analysis:
            print("❌ 没有结构分析结果")
            return
        
        analysis = self.structure_analysis
        print(f"\n📊 Excel文件结构分析摘要")
        print(f"📄 文件: {analysis['file_path']}")
        print(f"📋 工作表数量: {analysis['total_sheets']}")
        print(f"📋 工作表名称: {', '.join(analysis['sheet_names'])}")
        
        for sheet_name, sheet_analysis in analysis['sheets_analysis'].items():
            print(f"\n  📄 【{sheet_name}】")
            
            basic = sheet_analysis['basic_info']
            print(f"    📏 大小: {basic['max_row']}行 × {basic['max_column']}列")
            
            merged = sheet_analysis['merged_cells']
            if merged['count'] > 0:
                print(f"    🔗 合并单元格: {merged['count']}个 (表头:{merged['header_merges']}, 数据:{merged['data_merges']})")
            
            regions = sheet_analysis['data_regions']
            if regions:
                region = regions[0]
                print(f"    📊 数据密度: {region['density']:.1%}")
                print(f"    📍 数据范围: 第{region['min_row']}-{region['max_row']}行, 第{region['min_col']}-{region['max_col']}列")
            
            tables = sheet_analysis['table_structures']
            if tables:
                table = tables[0]
                header_info = table['header_analysis']
                print(f"    📋 建议表头行: 第{header_info['suggested_header_row']}行 (置信度:{header_info['confidence']:.1%})")
                print(f"    📊 表格类型: {table['table_type']}")
                
                quality = table.get('data_quality', {})
                if quality:
                    print(f"    ✅ 数据完整性: {quality['completeness']:.1%}")
            
            suggestions = sheet_analysis['read_suggestions']
            if suggestions.get('warnings'):
                print(f"    ⚠️  注意事项:")
                for warning in suggestions['warnings']:
                    print(f"        • {warning}")
    
    @staticmethod
    def read_excel_with_merged_cells_complete(file_path: str, sheet_name: str = None, max_rows: int = 1000) -> Tuple[pd.DataFrame, str]:
        """读取Excel文件并处理合并单元格，确保从第一行开始读取所有数据，不跳过任何内容"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            
            ws = wb[sheet_name]
            
            # 获取合并单元格信息
            merged_cells = ws.merged_cells.ranges
            merged_dict = {}
            
            # 创建合并单元格映射
            for merged_range in merged_cells:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_value = top_left_cell.value
                
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_dict[(row, col)] = top_left_value
            
            # 读取数据 - 从第一行开始，不跳过任何行
            data = []
            max_col = min(ws.max_column, 100)  # 限制最大列数
            max_row = min(ws.max_row, max_rows)
            
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    # 检查是否是合并单元格
                    if (row, col) in merged_dict:
                        value = merged_dict[(row, col)]
                    else:
                        cell = ws.cell(row=row, column=col)
                        value = cell.value
                    
                    row_data.append(value if value is not None else "")
                
                data.append(row_data)
            
            # 智能处理列名：尝试使用第一行作为字段名
            if data:
                if len(data) > 1:
                    # 检查第一行是否看起来像字段名
                    first_row = data[0]
                    looks_like_headers = True
                    
                    # 检查第一行是否主要包含文本且不全为空
                    non_empty_count = 0
                    text_count = 0
                    
                    for cell_value in first_row:
                        if cell_value is not None and str(cell_value).strip():
                            non_empty_count += 1
                            # 检查是否为文本（非纯数字）
                            try:
                                float(str(cell_value))
                            except (ValueError, TypeError):
                                text_count += 1
                    
                    # 如果第一行有足够的文本内容，认为是字段名
                    if non_empty_count > 0 and text_count / non_empty_count > 0.5:
                        # 使用第一行作为列名
                        columns = []
                        column_counts = {}
                        
                        for i, col_value in enumerate(first_row):
                            if col_value is None or str(col_value).strip() == "":
                                base_name = f"列{i+1}"
                            else:
                                base_name = str(col_value).strip()
                            
                            # 处理重复列名
                            if base_name in column_counts:
                                column_counts[base_name] += 1
                                unique_name = f"{base_name}_{column_counts[base_name]}"
                            else:
                                column_counts[base_name] = 0
                                unique_name = base_name
                            
                            columns.append(unique_name)
                        
                        # 从第二行开始作为数据
                        data_rows = data[1:] if len(data) > 1 else []
                        df = pd.DataFrame(data_rows, columns=columns)
                    else:
                        # 第一行不像字段名，生成默认列名并保留所有数据
                        columns = [f"列{i+1}" for i in range(len(data[0]))]
                        df = pd.DataFrame(data, columns=columns)
                else:
                    # 只有一行数据，生成默认列名
                    columns = [f"列{i+1}" for i in range(len(data[0]))]
                    df = pd.DataFrame(data, columns=columns)
            else:
                df = pd.DataFrame()
            
            # 清理数据 - 添加错误处理
            try:
                if not df.empty:
                    # 对每列进行数据类型规范化，避免混合类型
                    for col in df.columns:
                        try:
                            # 检查列是否包含混合类型
                            if df[col].dtype == 'object':
                                # 尝试将全部转换为字符串，避免混合类型
                                df[col] = df[col].astype(str)
                                # 将字符串'nan'和空字符串保持原样
                                df[col] = df[col].replace('nan', '')
                        except Exception as e:
                            print(f"列 {col} 数据类型转换警告: {e}")
                            # 如果转换失败，强制转换为字符串
                            df[col] = df[col].astype(str).replace('nan', '')
                    
                    # 应用infer_objects
                    df = df.infer_objects(copy=False)
            except Exception as e:
                print(f"数据清理时出现问题: {e}")
            
            # 生成markdown格式预览
            markdown_content = AdvancedExcelProcessor.df_to_markdown(df, sheet_name)
            
            return df, markdown_content
            
        except Exception as e:
            raise Exception(f"读取Excel文件时出错: {str(e)}")
    
    @staticmethod
    def read_excel_with_merged_cells(file_path: str, sheet_name: str = None, max_rows: int = 1000) -> Tuple[pd.DataFrame, str]:
        """读取Excel文件并处理合并单元格，支持更多行数"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            
            ws = wb[sheet_name]
            
            # 获取合并单元格信息
            merged_cells = ws.merged_cells.ranges
            merged_dict = {}
            
            # 创建合并单元格映射
            for merged_range in merged_cells:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_value = top_left_cell.value
                
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_dict[(row, col)] = top_left_value
            
            # 读取数据
            data = []
            max_col = min(ws.max_column, 100)  # 限制最大列数
            max_row = min(ws.max_row, max_rows)
            
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    # 检查是否是合并单元格
                    if (row, col) in merged_dict:
                        value = merged_dict[(row, col)]
                    else:
                        cell = ws.cell(row=row, column=col)
                        value = cell.value
                    
                    row_data.append(value if value is not None else "")
                
                data.append(row_data)
            
            # 转换为DataFrame
            if data and len(data) > 1:
                # 使用第一行作为列名，确保列名是字符串
                columns = []
                column_counts = {}  # 用于跟踪重复列名
                
                for i, col in enumerate(data[0]):
                    if col is None or str(col).strip() == "":
                        base_name = f"列{i+1}"
                    else:
                        base_name = str(col).strip()
                    
                    # 处理重复列名
                    if base_name in column_counts:
                        column_counts[base_name] += 1
                        unique_name = f"{base_name}_{column_counts[base_name]}"
                    else:
                        column_counts[base_name] = 0
                        unique_name = base_name
                    
                    columns.append(unique_name)
                
                # 创建DataFrame，确保数据行存在
                data_rows = data[1:] if len(data) > 1 else []
                if data_rows:
                    df = pd.DataFrame(data_rows, columns=columns)
                else:
                    df = pd.DataFrame(columns=columns)
            else:
                df = pd.DataFrame()
            
            # 清理数据 - 添加错误处理
            try:
                if not df.empty:
                    # 修复pandas FutureWarning并确保数据类型一致
                    df = df.replace('', np.nan)
                    
                    # 对每列进行数据类型规范化，避免混合类型
                    for col in df.columns:
                        try:
                            # 检查列是否包含混合类型
                            if df[col].dtype == 'object':
                                # 尝试将全部转换为字符串，避免混合类型
                                df[col] = df[col].astype(str)
                                # 将字符串'nan'重新替换为真正的NaN
                                df[col] = df[col].replace('nan', np.nan)
                        except Exception as e:
                            print(f"列 {col} 数据类型转换警告: {e}")
                            # 如果转换失败，强制转换为字符串
                            df[col] = df[col].astype(str).replace('nan', np.nan)
                    
                    # 最后应用infer_objects
                    df = df.infer_objects(copy=False)
                    
                    # 清理列名中的特殊字符
                    df.columns = [str(col).replace('\n', ' ').replace('\r', ' ').strip() for col in df.columns]
            except Exception as e:
                print(f"数据清理时出现问题: {e}")
                # 如果清理失败，至少确保基本的数据结构
                if not df.empty:
                    try:
                        df.columns = [str(col).replace('\n', ' ').replace('\r', ' ').strip() for col in df.columns]
                    except:
                        pass
            
            # 生成markdown格式预览
            markdown_content = AdvancedExcelProcessor.df_to_markdown(df, sheet_name)
            
            return df, markdown_content
            
        except Exception as e:
            raise Exception(f"读取Excel文件时出错: {str(e)}")
    
    @staticmethod
    def df_to_markdown(df: pd.DataFrame, sheet_name: str = "") -> str:
        """将DataFrame转换为Markdown格式，支持更好的格式化"""
        if df.empty:
            return f"## 📋 {sheet_name}\n\n*此工作表为空*\n\n"
        
        markdown = f"## 📋 {sheet_name}\n\n"
        
        # 数据基本信息
        markdown += f"**数据概览**: {len(df)} 行 × {len(df.columns)} 列\n\n"
        
        # 限制显示的行数和列数
        display_rows = min(20, len(df))
        display_cols = min(8, len(df.columns))
        display_df = df.head(display_rows).iloc[:, :display_cols]
        
        # 处理列名，确保不会太长
        headers = ["#"] + [str(col)[:15] + "..." if len(str(col)) > 15 else str(col) for col in display_df.columns]
        markdown += "| " + " | ".join(headers) + " |\n"
        markdown += "| " + " | ".join(["---"] * len(headers)) + " |\n"
        
        # 添加数据行
        for idx, row in display_df.iterrows():
            row_data = [str(idx + 1)]  # 行号从1开始
            for val in row:
                str_val = str(val) if pd.notna(val) else ""
                # 限制单元格内容长度
                if len(str_val) > 20:
                    str_val = str_val[:17] + "..."
                row_data.append(str_val)
            markdown += "| " + " | ".join(row_data) + " |\n"
        
        # 添加省略信息
        if len(df) > display_rows:
            markdown += f"\n*📝 仅显示前{display_rows}行，总共{len(df)}行*\n"
        
        if len(df.columns) > display_cols:
            markdown += f"*📊 仅显示前{display_cols}列，总共{len(df.columns)}列*\n"
        
        # 添加数据类型信息 - 修复错误处理
        markdown += "\n### 📈 数据类型概览\n"
        try:
            cols_to_show = list(df.columns)[:5]  # 只显示前5列的类型
            for col in cols_to_show:
                try:
                    if col in df.columns:
                        dtype = str(df[col].dtype)
                        non_null_count = df[col].count()
                        markdown += f"- **{col}**: {dtype} ({non_null_count}/{len(df)} 非空)\n"
                except Exception as e:
                    # 如果获取某列的数据类型失败，跳过该列
                    markdown += f"- **{col}**: 未知类型\n"
            
            if len(df.columns) > 5:
                markdown += f"- *... 还有 {len(df.columns) - 5} 列*\n"
        except Exception as e:
            markdown += "- *数据类型信息获取失败*\n"
        
        return markdown
    
    @staticmethod
    def df_to_pure_markdown(df: pd.DataFrame) -> str:
        """将DataFrame转换为纯净的Markdown表格格式，不包含任何分析信息"""
        if df.empty:
            return "*此工作表为空*\n"
        
        # 构建markdown表格
        markdown_lines = []
        
        # 判断是否使用第一行作为表头
        # 如果列名是"列1","列2"这样的通用名称，说明第一行是真正的表头数据
        use_first_row_as_header = all(col.startswith("列") and col[1:].isdigit() for col in df.columns)
        
        if use_first_row_as_header and len(df) > 0:
            # 使用第一行作为表头
            first_row = df.iloc[0]
            headers = []
            for val in first_row:
                if pd.isna(val) or str(val).strip() == "":
                    headers.append("")
                else:
                    header_str = str(val).replace("|", "\\|").replace("\n", " ").replace("\r", "").strip()
                    headers.append(header_str)
            
            markdown_lines.append("| " + " | ".join(headers) + " |")
            markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            
            # 从第二行开始作为数据行
            data_rows = df.iloc[1:]
        else:
            # 使用DataFrame的列名作为表头
            headers = [str(col) for col in df.columns]
            markdown_lines.append("| " + " | ".join(headers) + " |")
            markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            
            # 所有行都作为数据行
            data_rows = df
        
        # 数据行
        for idx, row in data_rows.iterrows():
            row_data = []
            for val in row:
                # 处理空值和格式化
                if pd.isna(val):
                    str_val = ""
                elif isinstance(val, float) and val.is_integer():
                    str_val = str(int(val))
                else:
                    str_val = str(val)
                
                # 处理特殊字符，确保markdown表格格式正确
                str_val = str_val.replace("|", "\\|").replace("\n", " ").replace("\r", "")
                row_data.append(str_val)
            
            markdown_lines.append("| " + " | ".join(row_data) + " |")
        
        return "\n".join(markdown_lines)
    
    def update_dataframe(self, sheet_name: str, df: pd.DataFrame):
        """更新指定工作表的数据"""
        self.modified_data[sheet_name] = df
    
    def add_calculated_column(self, sheet_name: str, column_name: str, formula_func, description: str = ""):
        """添加计算列"""
        if sheet_name in self.modified_data:
            df = self.modified_data[sheet_name]
            try:
                df[column_name] = formula_func(df)
                self.modified_data[sheet_name] = df
                return True, f"成功添加计算列: {column_name}"
            except Exception as e:
                return False, f"添加计算列失败: {str(e)}"
        return False, "工作表不存在"
    
    def fill_missing_values(self, sheet_name: str, column: str, method: str = "mean", custom_value=None):
        """填充缺失值"""
        if sheet_name not in self.modified_data:
            return False, "工作表不存在"
        
        df = self.modified_data[sheet_name]
        if column not in df.columns:
            return False, f"列 '{column}' 不存在"
        
        try:
            if method == "mean" and pd.api.types.is_numeric_dtype(df[column]):
                df[column].fillna(df[column].mean(), inplace=True)
            elif method == "median" and pd.api.types.is_numeric_dtype(df[column]):
                df[column].fillna(df[column].median(), inplace=True)
            elif method == "mode":
                mode_value = df[column].mode().iloc[0] if not df[column].mode().empty else ""
                df[column].fillna(mode_value, inplace=True)
            elif method == "forward":
                df[column].fillna(method='ffill', inplace=True)
            elif method == "backward":
                df[column].fillna(method='bfill', inplace=True)
            elif method == "custom" and custom_value is not None:
                df[column].fillna(custom_value, inplace=True)
            else:
                return False, f"不支持的填充方法: {method}"
            
            self.modified_data[sheet_name] = df
            return True, f"成功填充列 '{column}' 的缺失值"
            
        except Exception as e:
            return False, f"填充失败: {str(e)}"
    
    def add_summary_statistics(self, sheet_name: str, target_columns: List[str] = None):
        """添加汇总统计"""
        if sheet_name not in self.modified_data:
            return False, "工作表不存在"
        
        df = self.modified_data[sheet_name]
        
        try:
            if target_columns is None:
                target_columns = df.select_dtypes(include=[np.number]).columns.tolist()
            
            if not target_columns:
                return False, "没有找到数值列"
            
            # 创建汇总统计
            summary_data = []
            for col in target_columns:
                if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                    stats = {
                        '列名': col,
                        '计数': df[col].count(),
                        '均值': df[col].mean(),
                        '中位数': df[col].median(),
                        '标准差': df[col].std(),
                        '最小值': df[col].min(),
                        '最大值': df[col].max(),
                        '缺失值': df[col].isnull().sum()
                    }
                    summary_data.append(stats)
            
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_sheet_name = f"{sheet_name}_统计汇总"
                self.modified_data[summary_sheet_name] = summary_df
                return True, f"成功创建统计汇总工作表: {summary_sheet_name}"
            else:
                return False, "没有可统计的数值数据"
                
        except Exception as e:
            return False, f"创建统计汇总失败: {str(e)}"
    
    def export_to_excel(self, output_path: str = None) -> str:
        """导出修改后的数据到Excel文件"""
        try:
            if output_path is None:
                timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"modified_excel_{timestamp}.xlsx"
            
            # 创建新的工作簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 删除默认工作表
            
            for sheet_name, df in self.modified_data.items():
                ws = wb.create_sheet(title=sheet_name)
                
                # 写入列标题
                for col_idx, column in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                
                # 写入数据
                for row_idx, row in df.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx + 2, column=col_idx, value=value)
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(output_path)
            return output_path
            
        except Exception as e:
            raise Exception(f"导出Excel文件失败: {str(e)}")
    
    def get_data_preview(self, sheet_name: str) -> str:
        """获取数据预览"""
        if sheet_name in self.modified_data:
            df = self.modified_data[sheet_name]
            return self.df_to_markdown(df, sheet_name)
        return "工作表不存在"

class DataAnalyzer:
    """数据分析类"""
    
    @staticmethod
    def detect_data_types(df: pd.DataFrame) -> Dict[str, str]:
        """检测数据类型"""
        type_info = {}
        try:
            for col in df.columns:
                try:
                    if col in df.columns and not df[col].empty:
                        if pd.api.types.is_numeric_dtype(df[col]):
                            if df[col].dtype == 'int64':
                                type_info[col] = "整数"
                            else:
                                type_info[col] = "浮点数"
                        elif pd.api.types.is_datetime64_any_dtype(df[col]):
                            type_info[col] = "日期时间"
                        elif pd.api.types.is_bool_dtype(df[col]):
                            type_info[col] = "布尔值"
                        else:
                            type_info[col] = "文本"
                    else:
                        type_info[col] = "未知"
                except Exception as e:
                    type_info[col] = "检测失败"
        except Exception as e:
            print(f"数据类型检测出错: {e}")
        return type_info
    
    @staticmethod
    def find_duplicates(df: pd.DataFrame) -> pd.DataFrame:
        """查找重复行"""
        try:
            if df.empty:
                return pd.DataFrame()
            duplicates = df[df.duplicated(keep=False)]
            return duplicates
        except Exception as e:
            print(f"查找重复数据时出错: {e}")
            return pd.DataFrame()
    
    @staticmethod
    def get_missing_value_report(df: pd.DataFrame) -> pd.DataFrame:
        """获取缺失值报告"""
        missing_data = []
        try:
            for col in df.columns:
                try:
                    missing_count = df[col].isnull().sum()
                    missing_percent = (missing_count / len(df)) * 100 if len(df) > 0 else 0
                    missing_data.append({
                        '列名': str(col),
                        '缺失数量': missing_count,
                        '缺失百分比': f"{missing_percent:.2f}%",
                        '数据类型': str(df[col].dtype) if col in df.columns else "未知"
                    })
                except Exception as e:
                    missing_data.append({
                        '列名': str(col),
                        '缺失数量': 0,
                        '缺失百分比': "0.00%",
                        '数据类型': "检测失败"
                    })
        except Exception as e:
            print(f"获取缺失值报告时出错: {e}")
        
        return pd.DataFrame(missing_data)
    
    @staticmethod
    def detect_outliers(df: pd.DataFrame, method: str = "iqr") -> Dict[str, List]:
        """检测异常值"""
        outliers = {}
        try:
            numeric_columns = df.select_dtypes(include=[np.number]).columns
            
            for col in numeric_columns:
                try:
                    if method == "iqr":
                        Q1 = df[col].quantile(0.25)
                        Q3 = df[col].quantile(0.75)
                        IQR = Q3 - Q1
                        lower_bound = Q1 - 1.5 * IQR
                        upper_bound = Q3 + 1.5 * IQR
                        outlier_indices = df[(df[col] < lower_bound) | (df[col] > upper_bound)].index.tolist()
                        outliers[col] = outlier_indices
                    elif method == "zscore":
                        try:
                            from scipy import stats
                            z_scores = np.abs(stats.zscore(df[col].dropna()))
                            outlier_indices = df[col].dropna().index[z_scores > 3].tolist()
                            outliers[col] = outlier_indices
                        except ImportError:
                            # 如果scipy不可用，回退到IQR方法
                            Q1 = df[col].quantile(0.25)
                            Q3 = df[col].quantile(0.75)
                            IQR = Q3 - Q1
                            lower_bound = Q1 - 1.5 * IQR
                            upper_bound = Q3 + 1.5 * IQR
                            outlier_indices = df[(df[col] < lower_bound) | (df[col] > upper_bound)].index.tolist()
                            outliers[col] = outlier_indices
                except Exception as e:
                    print(f"检测列 {col} 的异常值时出错: {e}")
                    outliers[col] = []
        except Exception as e:
            print(f"异常值检测出错: {e}")
        
        return outliers 

class LightweightExcelAnalyzer:
    """轻量级Excel分析器 - 专为AI智能分析tab设计，生成精简提示词"""
    
    def __init__(self):
        self.workbook = None
        self.analysis_cache = None
    
    def quick_analyze(self, file_path: str) -> Dict[str, Any]:
        """快速分析Excel文件，生成AI分析所需的核心信息"""
        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            analysis = {
                'file_info': {
                    'filename': file_path.split('/')[-1],
                    'sheet_count': len(self.workbook.sheetnames),
                    'sheet_names': self.workbook.sheetnames
                },
                'sheets_summary': {},
                'ai_prompt': ""
            }
            
            # 分析每个工作表
            for sheet_name in self.workbook.sheetnames:
                sheet_summary = self._quick_analyze_sheet(sheet_name)
                analysis['sheets_summary'][sheet_name] = sheet_summary
            
            # 生成精简AI提示词
            analysis['ai_prompt'] = self._generate_compact_prompt(analysis)
            
            self.analysis_cache = analysis
            return analysis
            
        except Exception as e:
            return {
                'error': f"分析失败: {str(e)}",
                'file_info': {'filename': file_path.split('/')[-1]},
                'ai_prompt': f"文件 {file_path.split('/')[-1]} 分析失败: {str(e)}"
            }
    
    def _quick_analyze_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """快速分析单个工作表"""
        ws = self.workbook[sheet_name]
        
        # 基本信息
        basic_info = {
            'rows': ws.max_row,
            'columns': ws.max_column,
            'has_merged_cells': len(list(ws.merged_cells.ranges)) > 0,
            'merged_count': len(list(ws.merged_cells.ranges))
        }
        
        # 智能检测表头和数据区域
        header_info = self._smart_detect_headers(ws)
        
        # 提取关键字段信息
        fields_info = self._extract_key_fields(ws, header_info)
        
        # 数据样本
        data_samples = self._get_compact_samples(ws, header_info, limit=3)
        
        return {
            'basic': basic_info,
            'header_analysis': header_info,
            'fields': fields_info,
            'samples': data_samples,
            'sheet_type': self._classify_sheet_type(sheet_name, basic_info, fields_info)
        }
    
    def _smart_detect_headers(self, ws) -> Dict[str, Any]:
        """智能检测表头，特别处理复杂的合并单元格结构"""
        merged_ranges = list(ws.merged_cells.ranges)
        
        # 分析前10行，找出最可能的表头行
        header_candidates = []
        
        for row_num in range(1, min(11, ws.max_row + 1)):
            row_score = 0
            non_empty_count = 0
            text_count = 0
            
            for col in range(1, min(ws.max_column + 1, 50)):  # 限制检查前50列
                cell = ws.cell(row_num, col)
                if cell.value is not None:
                    non_empty_count += 1
                    
                    # 检查是否为文本（表头通常是文本）
                    try:
                        float(cell.value)
                    except (ValueError, TypeError):
                        text_count += 1
                        row_score += 1
                    
                    # 检查是否在合并单元格中（表头常有合并）
                    for merged_range in merged_ranges:
                        if (merged_range.min_row <= row_num <= merged_range.max_row and 
                            merged_range.min_col <= col <= merged_range.max_col):
                            row_score += 0.5
                            break
            
            if non_empty_count > 0:
                text_ratio = text_count / non_empty_count
                if text_ratio > 0.5 and non_empty_count >= 3:  # 至少3个有内容的单元格，且大部分是文本
                    header_candidates.append({
                        'row': row_num,
                        'score': row_score,
                        'non_empty': non_empty_count,
                        'text_ratio': text_ratio
                    })
        
        # 选择最佳表头行
        best_header_row = 1
        if header_candidates:
            # 综合评分：内容数量 + 文本比例 + 合并单元格加分
            best_candidate = max(header_candidates, 
                               key=lambda x: x['non_empty'] * x['text_ratio'] + x['score'])
            best_header_row = best_candidate['row']
        
        return {
            'suggested_header_row': best_header_row,
            'candidates': header_candidates,
            'has_complex_structure': len(merged_ranges) > 5,
            'merged_cell_analysis': self._analyze_header_merges(ws, merged_ranges, best_header_row)
        }
    
    def _analyze_header_merges(self, ws, merged_ranges: list, header_row: int) -> Dict[str, Any]:
        """分析表头区域的合并单元格"""
        header_merges = []
        
        # 查找表头附近的合并单元格
        for merged_range in merged_ranges:
            if abs(merged_range.min_row - header_row) <= 2:  # 表头前后2行内
                top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                merge_info = {
                    'range': str(merged_range),
                    'value': str(top_left_cell.value)[:30] if top_left_cell.value else "",
                    'span_cols': merged_range.max_col - merged_range.min_col + 1,
                    'span_rows': merged_range.max_row - merged_range.min_row + 1
                }
                header_merges.append(merge_info)
        
        return {
            'header_merges_count': len(header_merges),
            'header_merges': header_merges[:5],  # 只保留前5个
            'has_multi_level_headers': any(m['span_rows'] > 1 for m in header_merges)
        }
    
    def _extract_key_fields(self, ws, header_info: Dict[str, Any]) -> List[Dict[str, str]]:
        """提取关键字段信息（精简版）"""
        header_row = header_info['suggested_header_row']
        fields = []
        
        # 扫描表头行，提取字段名
        for col in range(1, min(ws.max_column + 1, 100)):  # 限制前100列
            cell = ws.cell(header_row, col)
            
            if cell.value is not None:
                field_name = str(cell.value).strip()
                
                # 如果字段名太短或为空，尝试向上查找（处理多层表头）
                if len(field_name) < 2:
                    for up_row in range(header_row - 1, max(0, header_row - 4), -1):
                        up_cell = ws.cell(up_row, col)
                        if up_cell.value and len(str(up_cell.value).strip()) > 1:
                            field_name = str(up_cell.value).strip()
                            break
                
                # 快速检测数据类型（只看前几个数据）
                data_type = self._quick_detect_type(ws, header_row + 1, col)
                
                fields.append({
                    'name': field_name,
                    'column': get_column_letter(col),
                    'type': data_type
                })
            
            # 如果连续遇到多个空列，可能已经超出数据范围
            elif col > 10:  # 前10列之后才开始这个判断
                empty_count = 0
                for check_col in range(col, min(col + 5, ws.max_column + 1)):
                    if ws.cell(header_row, check_col).value is None:
                        empty_count += 1
                if empty_count >= 5:  # 连续5个空列
                    break
        
        return fields
    
    def _quick_detect_type(self, ws, start_row: int, col: int) -> str:
        """快速检测字段类型"""
        sample_values = []
        
        # 只检查前10个数据
        for row in range(start_row, min(start_row + 10, ws.max_row + 1)):
            cell = ws.cell(row, col)
            if cell.value is not None:
                sample_values.append(cell.value)
        
        if not sample_values:
            return "空值"
        
        # 简单类型检测
        numeric_count = 0
        for value in sample_values:
            try:
                float(value)
                numeric_count += 1
            except (ValueError, TypeError):
                pass
        
        if numeric_count > len(sample_values) * 0.7:
            return "数值"
        else:
            return "文本"
    
    def _get_compact_samples(self, ws, header_info: Dict[str, Any], limit: int = 3) -> List[Dict[str, Any]]:
        """获取紧凑的数据样本"""
        header_row = header_info['suggested_header_row']
        samples = []
        
        # 获取字段名
        field_names = []
        for col in range(1, min(ws.max_column + 1, 20)):  # 只取前20列
            cell = ws.cell(header_row, col)
            field_name = str(cell.value).strip() if cell.value else f"列{col}"
            field_names.append(field_name)
        
        # 获取样本数据
        sample_count = 0
        for row in range(header_row + 1, ws.max_row + 1):
            if sample_count >= limit:
                break
            
            sample = {}
            has_data = False
            
            for i, col in enumerate(range(1, min(len(field_names) + 1, 21))):
                cell = ws.cell(row, col)
                value = cell.value
                
                if value is not None:
                    has_data = True
                    # 限制值的长度
                    if isinstance(value, str) and len(value) > 30:
                        value = value[:30] + "..."
                
                sample[field_names[i]] = value
            
            if has_data:
                samples.append(sample)
                sample_count += 1
        
        return samples
    
    def _classify_sheet_type(self, sheet_name: str, basic_info: Dict, fields_info: List) -> str:
        """分类工作表类型"""
        name_lower = sheet_name.lower()
        
        # 基于名称判断
        if any(keyword in name_lower for keyword in ['源数据', 'source', 'data', '明细']):
            return "源数据表"
        elif any(keyword in name_lower for keyword in ['汇总', 'summary', '总表', '统计']):
            return "汇总表"
        elif any(keyword in name_lower for keyword in ['模板', 'template', '配置']):
            return "模板表"
        elif any(keyword in name_lower for keyword in ['框架', 'framework', '结构']):
            return "框架表"
        
        # 基于数据特征判断
        rows, cols = basic_info['rows'], basic_info['columns']
        
        if rows > 1000:
            return "大数据表"
        elif cols > 50:
            return "宽表"
        elif rows < 50 and cols < 10:
            return "配置表"
        else:
            return "标准表"
    
    def _generate_compact_prompt(self, analysis: Dict[str, Any]) -> str:
        """生成精简的AI分析提示词"""
        file_info = analysis['file_info']
        sheets_summary = analysis['sheets_summary']
        
        prompt_parts = []
        
        # 文件概述
        prompt_parts.append(f"# Excel文件: {file_info['filename']}")
        prompt_parts.append(f"包含 {file_info['sheet_count']} 个工作表: {', '.join(file_info['sheet_names'])}")
        prompt_parts.append("")
        
        # 工作表概述
        prompt_parts.append("## 工作表概述")
        for sheet_name, summary in sheets_summary.items():
            basic = summary['basic']
            sheet_type = summary['sheet_type']
            fields_count = len(summary['fields'])
            
            prompt_parts.append(f"### {sheet_name} ({sheet_type})")
            prompt_parts.append(f"- 规模: {basic['rows']}行 × {basic['columns']}列")
            prompt_parts.append(f"- 字段数: {fields_count}")
            
            if basic['has_merged_cells']:
                prompt_parts.append(f"- ⚠️ 包含{basic['merged_count']}个合并单元格")
            
            # 关键字段（只显示前10个）
            key_fields = summary['fields'][:10]
            if key_fields:
                field_names = [f['name'] for f in key_fields if f['name'] and len(f['name']) > 1]
                if field_names:
                    prompt_parts.append(f"- 主要字段: {', '.join(field_names[:8])}")
            
            # 数据样本（只显示第一行）
            if summary['samples']:
                sample = summary['samples'][0]
                sample_items = []
                for k, v in list(sample.items())[:5]:  # 只显示前5个字段
                    if v is not None and str(v).strip():
                        sample_items.append(f"{k}={v}")
                if sample_items:
                    prompt_parts.append(f"- 样本: {', '.join(sample_items)}")
            
            prompt_parts.append("")
        
        # 分析建议
        prompt_parts.append("## 分析建议")
        
        # 找出可能的主表
        main_sheets = [name for name, summary in sheets_summary.items() 
                      if summary['sheet_type'] in ['源数据表', '大数据表']]
        if main_sheets:
            prompt_parts.append(f"- 主要数据源: {', '.join(main_sheets)}")
        
        # 识别关联关系
        all_fields = {}
        for sheet_name, summary in sheets_summary.items():
            field_names = [f['name'] for f in summary['fields'] if f['name'] and len(f['name']) > 1]
            all_fields[sheet_name] = field_names
        
        # 查找共同字段
        common_fields_found = False
        for sheet1, fields1 in all_fields.items():
            for sheet2, fields2 in all_fields.items():
                if sheet1 >= sheet2:
                    continue
                common = set(fields1) & set(fields2)
                if len(common) >= 2:
                    prompt_parts.append(f"- {sheet1} 与 {sheet2} 可通过字段关联: {', '.join(list(common)[:3])}")
                    common_fields_found = True
        
        if not common_fields_found:
            prompt_parts.append("- 未发现明显的表间关联字段")
        
        # 数据质量提醒
        complex_sheets = [name for name, summary in sheets_summary.items() 
                         if summary['basic']['has_merged_cells']]
        if complex_sheets:
            prompt_parts.append(f"- ⚠️ 注意: {', '.join(complex_sheets)} 存在合并单元格，处理时需特别注意")
        
        return "\n".join(prompt_parts)
    
    def get_analysis_cache(self) -> Dict[str, Any]:
        """获取缓存的分析结果"""
        return self.analysis_cache or {} 